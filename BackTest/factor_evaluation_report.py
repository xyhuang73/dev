# -*- coding: utf-8 -*-
"""
全因子评估 Excel 报告：参数页、指标说明、**胜率与收益口径**、全量结果、TopN 对比。

输出目录：项目根下 ``reports/factor/``。依赖：openpyxl（已在项目 requirements 中）。
"""
from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .factor_batch_job import FactorEvalRow, rank_rows_by_selection_rule, top_n_factors as take_top_n

# ---------------------------------------------------------------------------
# 报告用文字说明：与 factor_evaluation.json / single_factor_parameters.json / factors 下 alpha_*_parameters.json、selection_rule 元数据键、截面模块口径一致
# ---------------------------------------------------------------------------
_PARAM_REMARKS_BY_SUFFIX: dict[str, str] = {
    "min_rank_ic_mean": (
        "【IC 门槛】全样本 Rank IC 日序列的算术平均须 ≥ 此值；"
        "Rank IC 为每个交易日截面上因子秩与次日收益秩的 Pearson 相关。"
        "若为 0 表示仅要求非负（仍须 IC 为有限值）。"
    ),
    "min_ic_ir": (
        "【IC 门槛】IC_IR = RankIC日均 / RankIC日标准差（样本标准差）；须 ≥ 此值。"
        "反映 IC 稳定性；有效 IC 日过少或方差为 0 时 IC_IR 非有限，因子判为不可行。"
    ),
    "annualization_days": (
        "【年化】分层多空日收益序列换算年化夏普时用的交易日数；A 股常用 252。"
    ),
    "n_quantiles": (
        "【分层】每个交易日按因子值分位组数（如 10=十分位）；"
        "多头/空头端取最高/最低组内股票等权平均收益，差值为当日多空收益。"
    ),
    "min_names_per_day": (
        "【截面门槛】某一交易日截面上，因子与 forward_ret 同时有效的股票数 vn 须 ≥ 此值，"
        "该日才参与 Rank IC；低于则整日跳过。"
        "若 Excel 中 n_days_skip_min_names 与 n_trade_dates 相等且 min_daily_valid_names < 此值，"
        "说明门槛过高，应调低本参数或扩大 max_symbols/数据区间。"
    ),
    "max_symbols": (
        "【行情规模】参与面板回测的标的个数上限（按本地/板块列表截取）；"
        "与 min_names_per_day 共同影响截面是否足够大。"
    ),
    "alpha_prepare_max_workers": (
        "【计算·包内】见 Config/single_factor_parameters.json 的 alpha_prepare_max_workers：Alpha 包 prepare_data 内「特征与特征」线程并行度；"
        "1=顺序；>1=ThreadPoolExecutor。若自定义因子存在列间依赖须保持为 1。"
    ),
    "alpha158_ts_windows": (
        "【单因子】Alpha158 时序类特征的回看窗口（日）列表；与批量评估 IC 参数无关。"
    ),
    "factor_eval_parallel_cap": (
        "【计算·因子间】0=因子级并行线程数等于本批「前 N 个」因子个数；"
        ">0 时再取 min(N, 本值) 作为封顶，避免 N 极大时线程过多。"
    ),
    "market_data_source": (
        "【行情】xtdata=经 xtquant 拉取（需 miniQMT 就绪）；"
        "local_datadir=仅读 userdata_mini/datadir 下日线 DAT，不加载 xtquant。"
    ),
    "skip_qmt_process_check": (
        "【可选】为 true 时跳过 QMT 进程检测（自担风险）。"
    ),
    "qmt_process_name_substrings": (
        "【可选】判定 miniQMT 已启动时，在 tasklist 中匹配的进程名子串列表。"
    ),
}

_META_REMARKS: dict[str, str] = {
    "run_at": "本次任务开始写入报告时的本地时间（ISO）。",
    "start_date": "回测区间起始（界面/配置，YYYYMMDD）。",
    "end_date": "回测区间结束（YYYYMMDD）。",
    "n_bar_rows": "合并后的行情长表行数（所有标的×交易日 bar）。",
    "n_symbols": "行情长表中不同 vt_symbol 个数。",
    "factor_eval_front_n": "本次批量仅评估 inner_registry 中按顺序的前 N 个因子（界面 spinBox_factor_front_n）；未填表示未截断。",
    "factor_level_parallel_workers": "本批因子截面评估（IC/分层）阶段实际使用的线程池大小。",
    "factor_eval_unique_packs": "本批涉及的因子包名列表（去重）；每个包先顺序 prepare_data 一次再并行评因子。",
}


def _remark_for_flat_key(key: str) -> str:
    """扁平化后的 meta 键 → 中文备注（evaluation_config_ / selection_rule_ / single_factor_parameters_）。"""
    if key in _META_REMARKS:
        return _META_REMARKS[key]
    for prefix in ("evaluation_config_", "selection_rule_", "single_factor_parameters_"):
        if key.startswith(prefix):
            suf = key[len(prefix) :]
            return _PARAM_REMARKS_BY_SUFFIX.get(suf, "")
    return ""


def _result_column_remarks() -> list[tuple[str, str]]:
    """「全部因子」表各列含义，供单独工作表展示。"""
    return [
        ("factor_id", "InnerStrategy 注册表因子编号（F000001…）。"),
        ("label", "界面展示用标签。"),
        ("pack", "因子包：alpha_101 / alpha_158。"),
        ("feature", "该条记录对应的特征列名。"),
        ("error", "计算异常时的 Python 异常类型与信息；空表示未抛错。"),
        ("selection_feasible", "是否同时满足 RankIC 均值、IC_IR 与分层夏普有限性等「IC 门槛与多空夏普」规则。"),
        ("selection_objective", "可行时为分层多空夏普（年化口径）；不可行时通常为空白或 -inf 逻辑（表中已做空值处理）。"),
        ("selection_reason", "可行/不可行的文字说明（阈值比较或数据无效原因）。"),
        ("rank_ic_mean", "全样本有效交易日 Rank IC 的算术平均。"),
        ("rank_ic_std", "Rank IC 日序列的样本标准差。"),
        ("ic_ir", "rank_ic_mean / rank_ic_std；序列过短或方差为 0 时为非有限值。"),
        ("long_short_mean_daily", "有效交易日「分层多空日收益」的均值。"),
        ("long_short_vol_daily", "分层多空日收益的样本标准差。"),
        ("long_short_sharpe", "多空均值/多空波动 × sqrt(annualization_days)，为年化夏普形式。"),
        (
            "rank_ic_win_rate",
            "IC 胜率：在「有有效 Rank IC」的交易日里，Rank IC>0 的天数占比（0~1，×100% 即百分比）。"
            "表示因子方向与次日收益秩同向的比例。",
        ),
        (
            "long_short_win_rate",
            "多空胜率：在「有有效分层多空日收益」的交易日里，当日多空收益>0 的天数占比（0~1）。"
            "表示多头端相对空头端占优的交易日比例。",
        ),
        (
            "long_short_cumulative_return",
            "区间多空累计收益（复利）：对每日分层多空收益 r_t 按 ∏(1+r_t)-1 连乘，近似「每日调仓」下的累计结果。",
        ),
        (
            "long_short_annualized_return_approx",
            "年化收益（近似）：(1+多空日均收益)^252 - 1（252 来自配置 annualization_days），"
            "便于与直觉上的「年化」对照；与上一列复利累计不是同一口径。",
        ),
        ("n_ic_days", "实际参与 Rank IC 计算的交易日天数（已按日历日聚合截面）。"),
        ("n_ls_days", "实际参与分层多空计算的交易日天数。"),
        ("coverage_mean", "各交易日截面上「因子与收益同时有效」的样本比例之平均。"),
        ("n_trade_dates", "长表中参与循环的交易日个数（按 _trade_date 去重）。"),
        ("n_days_skip_min_names", "当日同时有效的因子与fwd_ret样本数 < min_names_per_day，被跳过 IC 的天数。若与 n_trade_dates 相等，说明阈值过高（见 min_daily_valid_names）。"),
        ("n_days_ic_nan_const_factor", "当日截面上因子值近似常数（标准差≈0），无法定义秩相关，RankIC 为 nan 的天数。"),
        ("n_days_ic_nan_const_fwd_ret", "当日截面上 fwd_ret 近似常数（例如收盘价对齐错误），RankIC 为 nan 的天数。"),
        ("n_days_ic_nan_small_cross", "截面有效样本数 < 3 的天数（一般仅当 min_names_per_day 配得很小时出现）。"),
        ("ic_unavailable_hint", "无 IC 日时的简短排障说明（与上列计数对照阅读）。"),
        ("min_daily_valid_names", "各交易日里 vn=「因子与fwd_ret同时有效」的个数；全样本取最小值，用于对照 min_names_per_day。"),
        ("median_daily_valid_names", "各交易日 vn 的中位数，反映典型截面宽度。"),
    ]


def _excel_scalar(v: Any) -> Any:
    """NaN/Inf 写入 openpyxl 常显示为空格，统一为可读占位。"""
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return ""
    return v


def _excel_pct(v: Any) -> str:
    """0~1 的胜率类小数 →「12.34%」便于扫表。"""
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return ""
    try:
        x = float(v)
    except (TypeError, ValueError):
        return ""
    return f"{x * 100.0:.2f}%"


def _excel_return_pct(v: Any) -> str:
    """收益率类小数（可正可负、可超 100%）→ 百分比字符串。"""
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return ""
    try:
        x = float(v)
    except (TypeError, ValueError):
        return ""
    return f"{x * 100.0:.2f}%"


def _flatten_dict(d: dict[str, Any], prefix: str = "") -> dict[str, Any]:
    """展开嵌套 dict；非标量叶子转为 str，便于写入单元格。"""
    out: dict[str, Any] = {}
    for k, v in d.items():
        key = f"{prefix}{k}" if not prefix else f"{prefix}_{k}"
        if isinstance(v, dict):
            out.update(_flatten_dict(v, key))
        elif isinstance(v, (str, int, float, bool)) or v is None:
            out[key] = v
        else:
            out[key] = str(v)
    return out


def _project_reports_dir() -> Path:
    """因子批量评估 Excel 输出目录：项目根下 ``reports/factor/``。"""
    root = Path(__file__).resolve().parent.parent
    p = root / "reports" / "factor"
    p.mkdir(parents=True, exist_ok=True)
    return p


def _win_rate_and_return_methodology_rows() -> list[tuple[str, str]]:
    """
    胜率与收益率的完整口径说明（与代码实现一致），写入独立工作表。

    对应实现::
        - ``fwd_ret``：``factor_alpha_runner.long_table_for_feature``
        - 截面 Rank IC / 分层多空 / 胜率 / 累计：``factor_cross_section_metrics``
    """
    return [
        (
            "一、收益 fwd_ret（次日持有收益）",
            "计算位置：对 raw_df 按 vt_symbol、datetime 排序后，"
            "fwd_ret = 下一根日线的收盘价 ÷ 当日收盘价 − 1（groupby 后 shift(-1) 的 close）。"
            "含义：在交易日 T 收盘买入、下一交易日 T+1 收盘卖出；持有期为相邻两根日线，不是「跳过一天自然日」也不是盘中价。"
            "同一交易日 T 的截面上，每只股票一行，因子值为 T 日可用值，收益为这段 T→T+1 的收益率，用于与因子对齐。",
        ),
        (
            "二、IC 胜率（rank_ic_win_rate）",
            "计算：在所有「该日 Rank IC 为有限值」的交易日上，统计 Rank IC > 0 的天数 ÷ 有效 IC 日数。"
            "Rank IC：当日截面上，因子值秩与 fwd_ret 秩的 Pearson 相关（见 _rank_ic_numpy）。"
            "含义：因子排名与「次日持有收益」排名同向的交易日比例；不是「明天上涨股票占比」，也不是单只股票涨跌次数。",
        ),
        (
            "三、多空胜率（long_short_win_rate）",
            "计算：在所有「该日分层多空收益为有限值」的交易日上，统计 当日多空收益 > 0 的天数 ÷ 有效分层日数。"
            "当日多空收益：按因子值将股票分为 n_quantiles 组，取因子最高一组与最低一组内股票的 fwd_ret 各自等权平均，二者之差（_decile_long_short_daily）。"
            "含义：高因子组相对低因子组在 T→T+1 持有期上占优的交易日比例。",
        ),
        (
            "四、区间累计收益（long_short_cumulative_return）",
            "计算：对有效分层日的日度多空收益序列 r_t，做连乘 ∏(1 + r_t) − 1。"
            "含义：把每日多空价差当作可逐日滚存的抽象组合收益（非单票实盘撮合）。",
        ),
        (
            "五、年化收益近似（long_short_annualized_return_approx）",
            "计算：(1 + 多空日收益算术均值) ^ annualization_days − 1；annualization_days 来自 Config/factor_evaluation.json（常为 252）。"
            "含义：由「日均多空收益」外推的年化倍数，与第四条的复利累计不是同一口径；便于与直觉上的年化对照。",
        ),
        (
            "六、为何有时胜率为空",
            "若 n_ic_days=0 或样本不足，IC 相关指标为空；若截面股票数不足以分层（见分层函数对最小样本的要求），"
            "n_ls_days 可能为 0，多空胜率与累计收益也会为空。可扩大 max_symbols、放宽 min_names_per_day 或拉长区间。",
        ),
    ]


def _write_methodology_sheet(ws: Any) -> None:  # noqa: ANN401 — openpyxl Worksheet
    """填充「胜率与收益口径」工作表并设置列宽与自动换行。"""
    ws.append(["项目", "说明（与程序实现一致）"])
    for title, body in _win_rate_and_return_methodology_rows():
        ws.append([title, body])
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 96
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.font = Font(bold=True)


def write_factor_evaluation_excel(
    rows: list[FactorEvalRow],
    meta: dict[str, Any],
    top_n: int = 10,
) -> Path:
    """
    写入工作簿：测试参数、指标列说明、**胜率与收益口径**、全部因子、TopN 对比。

    「胜率与收益口径」工作表中的定义与 ``factor_cross_section_metrics``、
    ``long_table_for_feature`` 代码一致。

    返回生成的 ``.xlsx`` 路径（位于项目根目录 ``reports/factor/`` 下）。
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = _project_reports_dir() / f"factor_evaluation_{ts}.xlsx"

    wb = Workbook()
    # --- 参数（键 / 值 / 备注）---
    ws0 = wb.active
    assert ws0 is not None
    ws0.title = "测试参数"
    flat_meta = _flatten_dict(dict(meta))
    ws0.append(["参数键", "当前值", "备注说明"])
    for k in sorted(flat_meta.keys()):
        ws0.append([k, flat_meta[k], _remark_for_flat_key(k)])
    ws0.column_dimensions["A"].width = 40
    ws0.column_dimensions["B"].width = 28
    ws0.column_dimensions["C"].width = 72

    # --- 指标与列说明（阅读报告时对照）---
    ws_help = wb.create_sheet("指标与列说明", 1)
    ws_help.append(["条目", "说明"])
    ws_help.append(
        [
            "阅读顺序",
            "各列含义见下方；「IC胜率、多空胜率、区间累计收益、年化近似」的算法与 fwd_ret 定义，"
            "以独立工作表「胜率与收益口径」为准（与代码一致）。",
        ],
    )
    ws_help.append(
        [
            "IC 门槛与多空夏普 · 总述",
            "先满足 RankIC 与 IC_IR 下限（IC 门槛），再以分层多空年化夏普作为主排序目标；"
            "若 IC 或分层序列为空（如 n_ic_days=0），指标多为非有限值，因子判为不可行。",
        ],
    )
    for name, desc in _result_column_remarks():
        ws_help.append([name, desc])
    ws_help.append(
        [
            "TopN_IC门槛与多空夏普 表",
            "与「全部因子」列含义一致；排序按 IC 门槛与多空夏普规则（可行优先，再按分层夏普）。"
            "优先看 IC胜率、多空胜率、区间累计收益、年化近似等直观列。",
        ],
    )
    ws_help.column_dimensions["A"].width = 28
    ws_help.column_dimensions["B"].width = 80
    for cell in ws_help[1]:
        cell.font = Font(bold=True)

    # --- 胜率与收益：定义与公式（与 factor_cross_section_metrics / long_table_for_feature 一致）---
    ws_method = wb.create_sheet("胜率与收益口径", 2)
    _write_methodology_sheet(ws_method)

    # --- 全量 ---
    ws1 = wb.create_sheet("全部因子")
    headers = [
        "factor_id",
        "label",
        "pack",
        "feature",
        "error",
        "selection_feasible",
        "selection_objective",
        "selection_reason",
        "rank_ic_mean",
        "rank_ic_std",
        "ic_ir",
        "long_short_mean_daily",
        "long_short_vol_daily",
        "long_short_sharpe",
        "IC胜率_RankIC为正日占比",
        "多空胜率_分层多空日收益为正日占比",
        "区间累计收益_多空复利",
        "年化收益_由日均近似",
        "n_ic_days",
        "n_ls_days",
        "coverage_mean",
        "n_trade_dates",
        "n_days_skip_min_names",
        "n_days_ic_nan_const_factor",
        "n_days_ic_nan_const_fwd_ret",
        "n_days_ic_nan_small_cross",
        "min_daily_valid_names",
        "median_daily_valid_names",
        "ic_unavailable_hint",
    ]
    ws1.append(headers)
    for r in rows:
        m = r.metrics
        ws1.append(
            [
                r.factor_id,
                r.label,
                r.pack,
                r.feature,
                r.error,
                r.selection_feasible,
                _excel_scalar(r.selection_objective),
                r.selection_reason,
                _excel_scalar(m.get("rank_ic_mean", "")),
                _excel_scalar(m.get("rank_ic_std", "")),
                _excel_scalar(m.get("ic_ir", "")),
                _excel_scalar(m.get("long_short_mean_daily", "")),
                _excel_scalar(m.get("long_short_vol_daily", "")),
                _excel_scalar(m.get("long_short_sharpe", "")),
                _excel_pct(m.get("rank_ic_win_rate", "")),
                _excel_pct(m.get("long_short_win_rate", "")),
                _excel_return_pct(m.get("long_short_cumulative_return", "")),
                _excel_return_pct(m.get("long_short_annualized_return_approx", "")),
                m.get("n_ic_days", ""),
                m.get("n_ls_days", ""),
                _excel_scalar(m.get("coverage_mean", "")),
                m.get("n_trade_dates", ""),
                m.get("n_days_skip_min_names", ""),
                m.get("n_days_ic_nan_const_factor", ""),
                m.get("n_days_ic_nan_const_fwd_ret", ""),
                m.get("n_days_ic_nan_small_cross", ""),
                m.get("min_daily_valid_names", ""),
                _excel_scalar(m.get("median_daily_valid_names", "")),
                m.get("ic_unavailable_hint", ""),
            ],
        )

    # --- TopN ---
    ws2 = wb.create_sheet(f"Top{top_n}_IC门槛与多空夏普")
    ranked = rank_rows_by_selection_rule(rows)
    top = take_top_n(ranked, n=top_n) if ranked else []
    ws2.append(
        [
            "排序",
            "factor_id",
            "label",
            "pack",
            "feature",
            "ic门槛满足",
            "主目标_分层多空夏普",
            "IC胜率",
            "多空胜率",
            "区间累计收益_多空复利",
            "年化收益_日均近似",
            "rank_ic_mean",
            "ic_ir",
            "long_short_sharpe",
            "error",
        ],
    )
    for i, r in enumerate(top, start=1):
        m = r.metrics
        ws2.append(
            [
                i,
                r.factor_id,
                r.label,
                r.pack,
                r.feature,
                r.selection_feasible,
                _excel_scalar(r.selection_objective),
                _excel_pct(m.get("rank_ic_win_rate", "")),
                _excel_pct(m.get("long_short_win_rate", "")),
                _excel_return_pct(m.get("long_short_cumulative_return", "")),
                _excel_return_pct(m.get("long_short_annualized_return_approx", "")),
                _excel_scalar(m.get("rank_ic_mean", "")),
                _excel_scalar(m.get("ic_ir", "")),
                _excel_scalar(m.get("long_short_sharpe", "")),
                r.error,
            ],
        )

    # 表头加粗（指标说明首行已单独加粗）
    for ws in (ws0, ws1, ws2):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    wb.save(path)
    return path
