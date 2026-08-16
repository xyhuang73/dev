# -*- coding: utf-8 -*-
"""
从 InnerStrategy/inner_registry.json 与源码生成「因子 / 策略」说明表，导出为 Excel 至 docs/。

用法（仓库根目录）::
    python scripts/export_registry_to_excel.py
"""
from __future__ import annotations

import ast
import json
import sys
from pathlib import Path

# 仓库根目录
_ROOT = Path(__file__).resolve().parent.parent
_INNER = _ROOT / "InnerStrategy"
_REGISTRY = _INNER / "inner_registry.json"
_DOCS = _ROOT / "docs"
_OUT_FILE = _DOCS / "因子与策略注册说明.xlsx"

# 因子包级说明（与业务文档一致，便于表格阅读）
_PACK_DOC: dict[str, str] = {
    "alpha_101": (
        "WorldQuant 风格 Alpha101：每条为一条表达式因子，由 VeighNa AlphaDataset 注册；"
        "回测时按编号定位到具体 feature 名。"
    ),
    "alpha_158": (
        "Qlib 风格 Alpha158：含 K 线结构、价格归一、多窗口时序等特征；"
        "部分行由源码中 for 循环按窗口展开为独立特征名。"
    ),
}


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise SystemExit(
            "请先安装: pip install openpyxl\n"
            f"原始错误: {exc}",
        ) from exc


def _load_registry() -> dict:
    if not _REGISTRY.is_file():
        raise SystemExit(f"未找到注册表: {_REGISTRY}，请先运行程序或 python -m InnerStrategy.inner_registry")
    return json.loads(_REGISTRY.read_text(encoding="utf-8"))


def _snippet_from_add_feature_line(line: str, feature: str, max_len: int = 450) -> str:
    """从源码行中截取 add_feature 第二参数字符串摘要。"""
    if f'add_feature("{feature}"' not in line and f"add_feature('{feature}'" not in line:
        return ""
    # 取第一个逗号后的内容，去掉末尾括号
    if "," not in line:
        return ""
    rest = line.split(",", 1)[1].strip()
    rest = rest.rstrip()
    if rest.endswith(")"):
        rest = rest[:-1].strip()
    if len(rest) > max_len:
        rest = rest[:max_len] + "…"
    return rest


def _find_factor_line_snippet(pack: str, feature: str) -> str:
    """在因子包 .py 中查找与 feature 对应的 add_feature 行并提取表达式摘要。"""
    fp = _INNER / "factors" / f"{pack}.py"
    if not fp.is_file():
        return ""
    for line in fp.read_text(encoding="utf-8").splitlines():
        if f'add_feature("{feature}"' in line or f"add_feature('{feature}'" in line:
            return _snippet_from_add_feature_line(line, feature)
    return ""


def _factor_row_description(pack: str, feature: str) -> str:
    """组合：包说明 + 特征说明 + 源码摘要（若有）。"""
    head = _PACK_DOC.get(pack, f"因子包「{pack}」，详见源码。")
    snip = _find_factor_line_snippet(pack, feature)
    if snip:
        return f"{head}\n【表达式摘要】{snip}"
    if pack == "alpha_158":
        return (
            f"{head}\n本行特征名「{feature}」多为循环展开得到；"
            f"请在 alpha_158.py 中按前缀（如 roc_/ma_/…）与窗口对照源码。"
        )
    return f"{head}\n特征「{feature}」：若未匹配到单行 add_feature，请打开 {pack}.py 搜索特征名。"


def _parse_strategy_docs(py_path: Path, class_name: str) -> tuple[str, str]:
    """返回 (模块文档字符串, 类文档字符串)。"""
    text = py_path.read_text(encoding="utf-8")
    tree = ast.parse(text)
    mod_doc = (ast.get_docstring(tree) or "").strip()
    cls_doc = ""
    for node in tree.body:
        if isinstance(node, ast.ClassDef) and node.name == class_name:
            cls_doc = (ast.get_docstring(node) or "").strip()
            break
    return mod_doc, cls_doc


def _strategy_row_description(module_stem: str, class_name: str) -> tuple[str, str, str]:
    """
    Returns:
        (模块文件路径说明, 模块级文档, 类级文档)
    """
    rel = f"InnerStrategy/strategies/{module_stem}.py"
    fp = _INNER / "strategies" / f"{module_stem}.py"
    if not fp.is_file():
        return rel, "", ""
    mod_doc, cls_doc = _parse_strategy_docs(fp, class_name)
    return rel, mod_doc, cls_doc


def export_excel() -> Path:
    """写入 docs/因子与策略注册说明.xlsx，返回输出路径。"""
    _ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    reg = _load_registry()
    _DOCS.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # --- 因子 ---
    ws_f = wb.active
    ws_f.title = "因子说明"
    headers_f = ["编号", "因子包", "特征名", "展示标签", "说明"]
    ws_f.append(headers_f)
    for cell in ws_f[1]:
        cell.font = Font(bold=True)
    for row in reg.get("factors", []):
        fid = row.get("id", "")
        pack = row.get("pack", "")
        feat = row.get("feature", "")
        label = row.get("label", "")
        desc = _factor_row_description(pack, feat)
        ws_f.append([fid, pack, feat, label, desc])
    for col in range(1, 6):
        letter = get_column_letter(col)
        ws_f.column_dimensions[letter].width = 18 if col <= 4 else 70
    for row in ws_f.iter_rows(min_row=2, max_row=ws_f.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- 策略 ---
    ws_s = wb.create_sheet("策略说明")
    headers_s = ["编号", "展示标签", "模块文件", "类名", "模块说明", "类说明", "相对路径"]
    ws_s.append(headers_s)
    for cell in ws_s[1]:
        cell.font = Font(bold=True)
    for row in reg.get("strategies", []):
        sid = row.get("id", "")
        mod = row.get("module", "")
        cls_name = row.get("class", "")
        label = row.get("label", "")
        path_note, mod_doc, cls_doc = _strategy_row_description(mod, cls_name)
        ws_s.append([sid, label, mod, cls_name, mod_doc, cls_doc, path_note])
    for col in range(1, 8):
        letter = get_column_letter(col)
        w = 14
        if col in (5, 6):
            w = 42
        if col == 7:
            w = 36
        ws_s.column_dimensions[letter].width = w
    for row in ws_s.iter_rows(min_row=2, max_row=ws_s.max_row, min_col=5, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(_OUT_FILE)
    return _OUT_FILE


def main() -> int:
    if str(_ROOT) not in sys.path:
        sys.path.insert(0, str(_ROOT))
    out = export_excel()
    print(f"已生成: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
